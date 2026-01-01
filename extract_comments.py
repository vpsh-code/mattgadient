#!/usr/bin/env python3
"""
mattgadient_comments_to_insights.py

Extract unstructured blog comments into a structured Excel "database" for analysis:
- Commenter + date + threading
- Rig/build extraction (motherboard/CPU/RAM/storage/HBA/NIC/PSU/platform)
- Power observations (watts) with sentence-bounded snippets and basic before/after cues
- Themes and insights
- Excel-safe output (no corruption) + presentable formatting

Run:
  python3 mattgadient_comments_to_insights.py \
    --url "https://mattgadient.com/7-watts-idle-on-intel-12th-13th-gen-the-foundation-for-building-a-low-power-server-nas/" \
    --out "mattgadient_low_power_comments_structured.xlsx"

Dependencies:
  pip install requests beautifulsoup4 lxml pandas openpyxl
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import requests
import pandas as pd
from bs4 import BeautifulSoup, Tag

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================
# Excel safety
# ============================
EXCEL_CELL_MAX = 32767
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def excel_safe_text(val: object, max_len: int = 32000) -> str:
    """Remove illegal XML chars and truncate to keep Excel stable."""
    if val is None:
        return ""
    s = str(val)
    s = _ILLEGAL_XML_RE.sub("", s)
    if len(s) > max_len:
        s = s[:max_len] + "…[TRUNCATED]"
    return s


# ============================
# Core regex patterns
# ============================
# Power / c-states
WATT_RE = re.compile(r"(?<!\w)(\d{1,4}(?:\.\d{1,2})?)\s*(w|watts?)\b", re.IGNORECASE)
CSTATE_RE = re.compile(r"\bC\s*([0-9]{1,2})\b", re.IGNORECASE)

# Context cues
METER_RE = re.compile(
    r"\b(kill[-\s]?a[-\s]?watt|watt\s*meter|wattmeter|smart\s*plug|from\s+the\s+wall|at\s+the\s+wall|wall\s+draw|wall\s+power|socket)\b",
    re.IGNORECASE,
)
IDLE_RE = re.compile(r"\b(idle|idling|at\s+idle|idles|standby|sleep|suspend)\b", re.IGNORECASE)
LOAD_RE = re.compile(r"\b(load|stress|prime95|furmark|full\s+load|100%\s*(load|cpu)|benchmark)\b", re.IGNORECASE)
BEFORE_RE = re.compile(r"\b(before|previously|used\s+to|was)\b", re.IGNORECASE)
AFTER_RE = re.compile(r"\b(after|now|currently|got\s+it\s+down\s+to|reduced\s+to|down\s+to)\b", re.IGNORECASE)

# Component/spec context (to avoid PSU ratings polluting idle)
PSU_CTX_RE = re.compile(r"\b(psu|power\s*supply|80\+|titanium|platinum|gold|seasonic|corsair|fsp|be\s*quiet)\b", re.IGNORECASE)
COMP_CTX_RE = re.compile(r"\b(nic|ethernet|rtl8125|i225|i226|hba|lsi|broadcom|sas|nvme|m\.2|ssd|hdd|drive|disk|controller|chipset|pcie)\b", re.IGNORECASE)
DRAW_RE = re.compile(r"\b(draw|drawing|consum(e|es|ed|ption)|uses|use|pulls)\b", re.IGNORECASE)

# Rig extraction (best-effort)
MB_BRAND_RE = re.compile(r"\b(ASRock|ASUS|Gigabyte|MSI|Supermicro)\b", re.IGNORECASE)
CPU_RE = re.compile(r"\b(i[3-9]-\d{4,5}[A-Z]*|Ryzen\s*\d{3,5}[A-Za-z]*|EPYC|Xeon)\b", re.IGNORECASE)
DDR_RE = re.compile(r"\bDDR[3-5]\b", re.IGNORECASE)
RAM_CAP_RE = re.compile(r"\b(\d{1,3})\s*GB\b", re.IGNORECASE)
RAM_KIT_RE = re.compile(r"\b(\d+)\s*x\s*(\d+)\s*GB\b", re.IGNORECASE)

# Storage patterns (quantities + sizes)
DRIVE_QTY_SIZE_RE = re.compile(r"\b(\d+)\s*[x×]\s*(\d+(?:\.\d+)?)\s*(TB|GB)\b", re.IGNORECASE)
NVME_RE = re.compile(r"\b(NVMe|M\.2|APST)\b", re.IGNORECASE)
SSD_RE = re.compile(r"\b(SSD|SATA\s*SSD)\b", re.IGNORECASE)
HDD_RE = re.compile(r"\b(HDD|hard\s*drive|spindown|spin\s*down|standby|APM)\b", re.IGNORECASE)

HBA_RE = re.compile(r"\b(HBA|SAS|LSI|Broadcom|SAS2008|SAS2308|SAS3008|9211-8i|9300-8i|9400|9500)\b", re.IGNORECASE)
NIC_RE = re.compile(r"\b(i219|i225|i226|Realtek|RTL8125|AQC|Aquantia|NIC|ethernet)\b", re.IGNORECASE)
PLATFORM_RE = re.compile(r"\b(Proxmox|ESXi|unRAID|TrueNAS|ZFS|KVM|VMware|Hyper-V|Linux|Windows)\b", re.IGNORECASE)
ASPM_RE = re.compile(r"\b(ASPM|L0s|L1\.?1|L1\.?2|Active\s*State\s*Power\s*Management)\b", re.IGNORECASE)
BIOS_RE = re.compile(r"\b(BIOS|UEFI|firmware)\b", re.IGNORECASE)

# Themes (analyst-friendly categories)
THEME_RULES: Dict[str, re.Pattern] = {
    "ASPM_PCIe_PowerMgmt": ASPM_RE,
    "BIOS_Firmware_Settings": re.compile(r"\b(BIOS|UEFI|firmware|c-state|C\s*\d{1,2}|S0ix)\b", re.IGNORECASE),
    "HBA_Controller_Impact": HBA_RE,
    "NIC_Driver_Offload_Power": re.compile(r"\b(i219|i225|i226|Realtek|RTL8125|Aquantia|AQC|NIC|ethernet|offload|TSO|GSO|GRO)\b", re.IGNORECASE),
    "NVMe_M2_Idle_Draw": re.compile(r"\b(NVMe|M\.2|APST)\b", re.IGNORECASE),
    "Storage_HDD_Spindown": HDD_RE,
    "PSU_Efficiency_Idle": re.compile(r"\b(PSU|power\s*supply|80\+|Titanium|Platinum|Gold)\b", re.IGNORECASE),
    "OS_Kernel_Tuning": re.compile(r"\b(powertop|tlp|kernel|linux|driver|module)\b", re.IGNORECASE),
    "Measurement_Methodology": re.compile(r"\b(kill[-\s]?a[-\s]?watt|meter|measured|measurement|wattmeter|wall|socket|smart\s*plug)\b", re.IGNORECASE),
}


# ============================
# Data model
# ============================
@dataclass
class CommentRecord:
    comment_id: int
    parent_comment_id: Optional[int]
    thread_depth: int
    commenter: str
    date_raw: str
    date_iso: Optional[str]       # YYYY-MM-DD
    date_ddmmyyyy: Optional[str]  # DD/MM/YYYY
    comment_text: str


# ============================
# Fetch
# ============================
def fetch_html(url: str, timeout: int = 30) -> str:
    headers = {"User-Agent": "Mozilla/5.0 (comments-to-insights)"}
    r = requests.get(url, headers=headers, timeout=timeout)
    r.raise_for_status()
    return r.text


# ============================
# Parsing helpers
# ============================
def parse_date_any(s: str) -> Tuple[str, Optional[str], Optional[str]]:
    raw = (s or "").strip()
    if not raw:
        return "", None, None

    raw2 = re.sub(r"\s+at\s+\d{1,2}:\d{2}\s*(am|pm)?\b.*$", "", raw, flags=re.IGNORECASE).strip()

    for fmt in ("%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw2, fmt).date()
            return raw, d.isoformat(), d.strftime("%d/%m/%Y")
        except ValueError:
            continue

    return raw, None, None


def clean_commenter(s: str) -> str:
    x = (s or "").strip()
    x = re.sub(r"\s*-\s*click\s+here\s+to\s+reply\s*-?\s*$", "", x, flags=re.IGNORECASE).strip()
    x = re.sub(r"\s+on\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}.*$", "", x, flags=re.IGNORECASE).strip()
    return x or "Unknown"


def wp_comment_id(li: Tag) -> Optional[int]:
    """
    WordPress commonly uses: <li id="comment-1234">.
    Return 1234 if present; else None.
    """
    cid = li.get("id", "") or ""
    m = re.match(r"comment-(\d+)$", cid.strip())
    if m:
        return int(m.group(1))
    return None


def own_comment_container(li: Tag) -> Tag:
    """
    Return the container representing THIS comment (not its children).
    Strategy:
      1) Prefer direct child <article> (WP typical).
      2) Else direct child with id like div-comment-XXXX.
      3) Else fall back to li itself (but we will remove children before extraction).
    """
    # Prefer direct article
    art = li.find("article", recursive=False)
    if art:
        return art

    # Sometimes it's a div with id="div-comment-XXXX"
    direct_divs = li.find_all("div", recursive=False)
    for d in direct_divs:
        did = d.get("id", "") or ""
        if did.startswith("div-comment-"):
            return d

    return li


def strip_non_content(container: Tag) -> None:
    """
    Remove metadata/header/reply/children blocks so they cannot leak into extracted text.
    Operates in-place.
    """
    # Remove child thread entirely from this comment's extraction
    for sel in ["ol.children", "ul.children"]:
        for n in container.select(sel):
            n.decompose()

    # Remove meta blocks and reply links if they exist within the container
    for sel in [".comment-meta", ".comment-metadata", ".comment-author", ".reply", ".comment-reply-link"]:
        for n in container.select(sel):
            n.decompose()


def extract_commenter_date(li: Tag) -> Tuple[str, str, Optional[str], Optional[str]]:
    """
    Markup-first; fallback parses header-style text:
      "Anonymous on May 14, 2023 - click here to reply"
    Critically: operate only on THIS comment's container (exclude children).
    """
    container = own_comment_container(li)

    # Work on a copy-ish approach: remove nested children/meta from a cloned fragment
    # BS4 doesn't provide a clean clone API; do minimal by avoiding child lists and
    # querying within the container first.
    commenter = ""
    date_raw = ""

    # Markup: commenter
    a = container.select_one(".comment-author .fn") or container.select_one(".fn")
    if a:
        commenter = a.get_text(" ", strip=True)

    # Markup: date
    t = container.select_one("time")
    if t:
        date_raw = t.get_text(" ", strip=True) or (t.get("datetime") or "")
    else:
        m = container.select_one(".comment-metadata a") or container.select_one(".comment-meta a")
        if m:
            date_raw = m.get_text(" ", strip=True)

    # Fallback: parse combined header text (container-only)
    if not commenter or not date_raw:
        header_text = ""
        header = container.select_one(".comment-meta") or container.select_one(".comment-metadata") or container.select_one(".comment-author")
        if header:
            header_text = header.get_text(" ", strip=True)
        else:
            header_text = container.get_text(" ", strip=True)[:250]

        mm = re.search(r"^(.*?)\s+on\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})\b", header_text)
        if mm:
            if not commenter:
                commenter = mm.group(1).strip()
            if not date_raw:
                date_raw = mm.group(2).strip()

    commenter = clean_commenter(commenter)
    dr, di, dmy = parse_date_any(date_raw)
    return commenter, dr, di, dmy


def extract_comment_body(li: Tag) -> str:
    """
    Extract THIS comment's body only (exclude nested child replies and meta headers).
    """
    container = own_comment_container(li)

    # Create a working copy by re-parsing this container's HTML so we can safely decompose
    # without mutating the original soup tree.
    frag = BeautifulSoup(str(container), "lxml")
    work = frag.body or frag  # body exists in lxml wrapper
    # Choose the first real element under body if present
    if getattr(work, "find", None):
        first_el = work.find(True)
        if first_el:
            work = first_el

    strip_non_content(work)

    # Prefer comment-content if available, but within this comment only
    body = work.select_one(".comment-content")
    if not body:
        # Some themes use "comment-content" alternatives; fall back to paragraphs
        body = work

    # Collect text from block-ish nodes, avoiding duplication
    parts: List[str] = []
    for node in body.find_all(["p", "pre", "code", "blockquote"], recursive=True):
        txt = node.get_text(" ", strip=True)
        if txt:
            parts.append(txt)

    text = "\n".join([p for p in parts if p]).strip()

    # If no p/pre/etc existed, fall back to stripped full text
    if not text:
        text = body.get_text("\n", strip=True).strip()

    return text


def extract_top_level_comments(soup: BeautifulSoup) -> List[Tag]:
    root = (
        soup.select_one("ol.comment-list")
        or soup.select_one("ol.commentlist")
        or soup.select_one("div#comments")
        or soup.select_one("#comments")
    )
    if not root:
        raise RuntimeError("Could not find comments container (#comments / comment-list).")

    if root.name in ("ol", "ul"):
        top_level = root.find_all("li", recursive=False)
    else:
        top_level = root.select("li.comment") or root.find_all("li")

    filtered = []
    for li in top_level:
        # Keep LI elements that look like comments
        if li.select_one("time") or li.select_one(".fn") or li.select_one(".comment-content") or li.select_one(".comment-metadata"):
            filtered.append(li)
    return filtered or top_level


def walk_comments(top_level_lis: List[Tag]) -> List[CommentRecord]:
    """
    Walk WP threaded comments.
    Uses WordPress comment IDs when available; otherwise falls back to an internal counter.
    Parent IDs are the WP IDs (when available) so joins are stable across runs.
    """
    seq_id = 0
    records: List[CommentRecord] = []

    def walk(li: Tag, parent: Optional[int], depth: int) -> None:
        nonlocal seq_id
        wp_id = wp_comment_id(li)
        if wp_id is None:
            seq_id += 1
            comment_id = seq_id
        else:
            comment_id = wp_id

        commenter, dr, di, dmy = extract_commenter_date(li)
        text = extract_comment_body(li)

        records.append(
            CommentRecord(
                comment_id=comment_id,
                parent_comment_id=parent,
                thread_depth=depth,
                commenter=commenter,
                date_raw=dr,
                date_iso=di,
                date_ddmmyyyy=dmy,
                comment_text=text,
            )
        )

        children = li.select_one("ol.children") or li.select_one("ul.children")
        if children:
            for child_li in children.find_all("li", recursive=False):
                # Parent should be THIS comment's ID (wp id if present)
                walk(child_li, parent=comment_id, depth=depth + 1)

    for li in top_level_lis:
        walk(li, parent=None, depth=0)

    return records


# ============================
# Sentence-bounded snippet for measurements
# ============================
SENT_BOUNDARY_RE = re.compile(r"(?<=[.!?])\s+|\n+")

def sentence_bounded_snippet(text: str, match_start: int, match_end: int, max_chars: int = 650) -> str:
    """
    Expand a watt match to sentence boundaries, and ensure the snippet ends cleanly.
    - Start: previous sentence boundary
    - End: next sentence boundary (optionally include one extra sentence if short)
    - If we must truncate, truncate at a sentence boundary (or add ellipsis at the end).
    """
    t = (text or "").strip()
    if not t:
        return ""

    # Find start boundary (previous punctuation boundary)
    start = 0
    for i in range(match_start, -1, -1):
        if t[i] in ".!?\n":
            start = i + 1
            break

    # Find end boundary (next punctuation boundary)
    end = len(t)
    for i in range(match_end, len(t)):
        if t[i] in ".!?\n":
            end = i + 1
            break

    snippet = t[start:end].strip()

    # If snippet is very short, include one more sentence (if available)
    if len(snippet) < 220 and end < len(t):
        for i in range(end, len(t)):
            if t[i] in ".!?\n":
                end2 = i + 1
                snippet = t[start:end2].strip()
                break

    # Hard cap while keeping sentence integrity
    if len(snippet) > max_chars:
        cut = snippet[:max_chars].rstrip()
        # Try to cut at the last sentence boundary within the cap
        last = max(cut.rfind("."), cut.rfind("!"), cut.rfind("?"))
        if last > 80:  # only if we have enough content
            cut = cut[: last + 1].rstrip()
            snippet = cut + " …"
        else:
            snippet = cut + " …"

    return snippet


# ============================
# Measurement context classifier
# ============================
def measurement_context(snippet: str, watts: float) -> Tuple[str, str, Optional[str]]:
    """
    Returns: (Context_Class, Confidence, Change_Direction)
    Change_Direction: "Before" / "After" / None (heuristic)
    """
    s = snippet or ""
    s_low = s.lower()

    # PSU rating/spec
    if PSU_CTX_RE.search(s) and (watts >= 200 or re.search(r"\b\d{3,4}\s*w\b", s_low)):
        ctx = "PSU_Rating/Spec"
        conf = "High"
        return ctx, conf, None

    # Load
    if LOAD_RE.search(s):
        ctx = "System_Load_Wall" if (METER_RE.search(s) or "wall" in s_low or "socket" in s_low) else "System_Load_Unknown"
        conf = "High" if METER_RE.search(s) else "Medium"
        return ctx, conf, ("After" if AFTER_RE.search(s) else ("Before" if BEFORE_RE.search(s) else None))

    # Idle
    if IDLE_RE.search(s):
        ctx = "System_Idle_Wall" if (METER_RE.search(s) or "wall" in s_low or "socket" in s_low) else "System_Idle_Unknown"
        conf = "High" if METER_RE.search(s) else "Medium"
        return ctx, conf, ("After" if AFTER_RE.search(s) else ("Before" if BEFORE_RE.search(s) else None))

    # Component draw
    if COMP_CTX_RE.search(s) and watts <= 30 and (DRAW_RE.search(s) or re.search(r"\bdraws\b|\bpulls\b|\buses\b", s_low)):
        return "Component_Draw", "Medium", None

    return "Unclear", "Low", None


# ============================
# Rig/build extraction (best-effort)
# ============================
def extract_rig(text: str) -> Dict[str, object]:
    t = text or ""
    out: Dict[str, object] = {}

    mb_brand = None
    m = MB_BRAND_RE.search(t)
    if m:
        mb_brand = m.group(1)
    out["MB_Brand"] = mb_brand or ""

    mb_model = ""
    if mb_brand:
        idx = t.lower().find(mb_brand.lower())
        window = t[idx: idx + 140] if idx >= 0 else t
        mm = re.search(r"\b(Z\d{3,4}|B\d{3,4}|H\d{3,4}|X\d{3,4})\b[^\n]{0,80}", window, flags=re.IGNORECASE)
        if mm:
            mb_model = mm.group(0).strip()
    out["MB_Model_Raw"] = mb_model

    cpu_hits = sorted(set([m.group(0) for m in CPU_RE.finditer(t)]))
    out["CPU_Models"] = "; ".join(cpu_hits)

    ddr = sorted(set([m.group(0).upper() for m in DDR_RE.finditer(t)]))
    out["DDR"] = "; ".join(ddr)

    kits = []
    for m in RAM_KIT_RE.finditer(t):
        kits.append(f"{m.group(1)}x{m.group(2)}GB")
    out["RAM_Kits"] = "; ".join(sorted(set(kits)))

    ram_caps = [int(m.group(1)) for m in RAM_CAP_RE.finditer(t)]
    out["RAM_GB_Mentions"] = "; ".join(str(x) for x in sorted(set(ram_caps)))

    qty_size = []
    for m in DRIVE_QTY_SIZE_RE.finditer(t):
        qty = int(m.group(1))
        size = float(m.group(2))
        unit = m.group(3).upper()
        qty_size.append(f"{qty}x{size:g}{unit}")
    out["Drive_Qty_Size_Mentions"] = "; ".join(sorted(set(qty_size)))

    out["Mentions_NVMe"] = bool(NVME_RE.search(t))
    out["Mentions_SSD"] = bool(SSD_RE.search(t))
    out["Mentions_HDD"] = bool(HDD_RE.search(t))

    out["Mentions_HBA"] = bool(HBA_RE.search(t))
    out["Mentions_NIC"] = bool(NIC_RE.search(t))
    out["Platforms"] = "; ".join(sorted(set([m.group(0) for m in PLATFORM_RE.finditer(t)])))

    out["Mentions_BIOS"] = bool(BIOS_RE.search(t))
    out["Mentions_ASPM"] = bool(ASPM_RE.search(t))

    return out


# ============================
# Themes
# ============================
def classify_themes(text: str) -> List[str]:
    hits = []
    for theme, rx in THEME_RULES.items():
        if rx.search(text or ""):
            hits.append(theme)
    return hits


# ============================
# Workbook formatting
# ============================
def style_sheet(ws, wrap_cols: Optional[List[str]] = None, widths: Optional[Dict[str, float]] = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    if wrap_cols:
        header = [c.value for c in ws[1]]
        idx_map = {header[i]: i + 1 for i in range(len(header))}
        for name in wrap_cols:
            idx = idx_map.get(name)
            if not idx:
                continue
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    if widths:
        header_row = [c.value for c in ws[1]]
        col_idx = {header_row[i]: i + 1 for i in range(len(header_row))}
        for col_name, w in widths.items():
            if col_name in col_idx:
                ws.column_dimensions[get_column_letter(col_idx[col_name])].width = w


def finalize_workbook(path: str) -> None:
    wb = load_workbook(path)

    if "README" in wb.sheetnames:
        ws = wb["README"]
        style_sheet(ws, wrap_cols=["Value"], widths={"Field": 26, "Value": 110})

    if "Comments" in wb.sheetnames:
        ws = wb["Comments"]
        style_sheet(
            ws,
            wrap_cols=["Comment_Text"],
            widths={
                "Comment_ID": 10,
                "Parent_Comment_ID": 16,
                "Thread_Depth": 12,
                "Commenter": 22,
                "Date_DDMMYYYY": 14,
                "Date_ISO": 12,
                "Comment_Text": 70,
            },
        )

    if "Rigs" in wb.sheetnames:
        ws = wb["Rigs"]
        style_sheet(ws, wrap_cols=["Rig_Notes"], widths={
            "Comment_ID": 10,
            "Commenter": 22,
            "Date_DDMMYYYY": 14,
            "MB_Brand": 14,
            "MB_Model_Raw": 30,
            "CPU_Models": 26,
            "DDR": 10,
            "RAM_Kits": 14,
            "RAM_GB_Mentions": 16,
            "Drive_Qty_Size_Mentions": 24,
            "Platforms": 18,
            "Rig_Notes": 60,
        })

    if "Power_Observations" in wb.sheetnames:
        ws = wb["Power_Observations"]
        style_sheet(ws, wrap_cols=["Sentence_Snippet"], widths={
            "Comment_ID": 10,
            "Watts": 10,
            "Context_Class": 18,
            "Confidence": 10,
            "Change_Cue": 10,
            "Mentions_Meter": 14,
            "Mentions_Idle": 13,
            "Mentions_Load": 13,
            "Sentence_Snippet": 90,
        })

    if "Themes" in wb.sheetnames:
        ws = wb["Themes"]
        style_sheet(ws, wrap_cols=["Comment_Text"], widths={
            "Comment_ID": 10,
            "Commenter": 22,
            "Date_DDMMYYYY": 14,
            "Theme": 28,
            "Comment_Text": 90,
        })

    if "Insights" in wb.sheetnames:
        ws = wb["Insights"]
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24

    if "QA_Sample" in wb.sheetnames:
        ws = wb["QA_Sample"]
        style_sheet(ws, wrap_cols=["Comment_Text", "Reviewer_Notes"], widths={
            "Comment_ID": 10,
            "Commenter": 22,
            "Date_DDMMYYYY": 14,
            "Evidence_Score_1to5": 18,
            "System_Idle_Wall_Min_W": 20,
            "Comment_Text": 80,
            "Reviewer_Notes": 40,
        })

    wb.save(path)


# ============================
# Build tables
# ============================
def build_tables(records: List[CommentRecord]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame]]:
    # COMMENTS
    comments_rows = []
    for r in records:
        comments_rows.append({
            "Comment_ID": r.comment_id,
            "Parent_Comment_ID": r.parent_comment_id,
            "Thread_Depth": r.thread_depth,
            "Commenter": excel_safe_text(r.commenter, max_len=200),
            "Date_Raw": excel_safe_text(r.date_raw, max_len=200),
            "Date_ISO": r.date_iso or "",
            "Date_DDMMYYYY": r.date_ddmmyyyy or "",
            "Comment_Text": excel_safe_text(r.comment_text, max_len=32000),
        })
    df_comments = pd.DataFrame(comments_rows)

    # POWER OBSERVATIONS
    pow_rows = []
    for r in records:
        t = r.comment_text or ""
        for m in WATT_RE.finditer(t):
            watts = float(m.group(1))
            snippet = sentence_bounded_snippet(t, m.start(), m.end(), max_chars=650)
            ctx, conf, change = measurement_context(snippet, watts)

            pow_rows.append({
                "Comment_ID": r.comment_id,
                "Watts": watts,
                "Context_Class": ctx,
                "Confidence": conf,
                "Change_Cue": change or "",
                "Mentions_Meter": bool(METER_RE.search(snippet)),
                "Mentions_Idle": bool(IDLE_RE.search(snippet)),
                "Mentions_Load": bool(LOAD_RE.search(snippet)),
                "Sentence_Snippet": excel_safe_text(snippet, max_len=4000),
            })
    df_power = pd.DataFrame(pow_rows)

    # THEMES (include FULL comment so it's actually useful)
    theme_rows = []
    for r in records:
        themes = classify_themes(r.comment_text)
        if not themes:
            continue
        for th in themes:
            theme_rows.append({
                "Comment_ID": r.comment_id,
                "Commenter": r.commenter,
                "Date_DDMMYYYY": r.date_ddmmyyyy or "",
                "Theme": th,
                "Comment_Text": excel_safe_text(r.comment_text, max_len=32000),
            })
    df_themes = pd.DataFrame(theme_rows)

    # FEATURES / EVIDENCE SCORE
    feat_rows = []
    for r in records:
        t = r.comment_text or ""
        cstates = sorted({int(x) for x in CSTATE_RE.findall(t) if x.isdigit()})

        idle_wall = df_power[
            (df_power["Comment_ID"] == r.comment_id) &
            (df_power["Context_Class"] == "System_Idle_Wall")
        ]["Watts"].tolist()

        evidence = 0
        if WATT_RE.search(t):
            evidence += 1
        if not df_power.empty and any(df_power[df_power["Comment_ID"] == r.comment_id]["Context_Class"].isin(["System_Idle_Wall", "System_Load_Wall"])):
            evidence += 2
        if METER_RE.search(t):
            evidence += 1
        if MB_BRAND_RE.search(t) or CPU_RE.search(t) or DRIVE_QTY_SIZE_RE.search(t):
            evidence += 1
        evidence = min(evidence, 5)

        feat_rows.append({
            "Comment_ID": r.comment_id,
            "Evidence_Score_1to5": evidence,
            "Word_Count": len(re.findall(r"\w+", t)),
            "CState_Max": max(cstates) if cstates else None,
            "CStates_Mentioned": "; ".join([f"C{x}" for x in cstates]),
            "Mentions_ASPM": bool(ASPM_RE.search(t)),
            "Mentions_BIOS": bool(BIOS_RE.search(t)),
            "Mentions_HBA": bool(HBA_RE.search(t)),
            "Mentions_NIC": bool(NIC_RE.search(t)),
            "Mentions_NVMe": bool(NVME_RE.search(t)),
            "Mentions_HDD": bool(HDD_RE.search(t)),
            "System_Idle_Wall_Min_W": min(idle_wall) if idle_wall else None,
        })
    df_features = pd.DataFrame(feat_rows)

    # RIGS
    rig_rows = []
    for r in records:
        rig = extract_rig(r.comment_text)
        signal = (
            bool(rig.get("MB_Brand"))
            or bool(rig.get("CPU_Models"))
            or bool(rig.get("Drive_Qty_Size_Mentions"))
            or rig.get("Mentions_HBA")
            or rig.get("Mentions_NVMe")
            or rig.get("Mentions_HDD")
        )
        if not signal:
            continue

        notes_bits = []
        if rig.get("Mentions_ASPM"):
            notes_bits.append("ASPM mentioned")
        if rig.get("Mentions_BIOS"):
            notes_bits.append("BIOS mentioned")
        if rig.get("Mentions_HBA"):
            notes_bits.append("HBA mentioned")
        if rig.get("Mentions_NIC"):
            notes_bits.append("NIC mentioned")
        rig_notes = "; ".join(notes_bits)

        rig_rows.append({
            "Comment_ID": r.comment_id,
            "Commenter": r.commenter,
            "Date_DDMMYYYY": r.date_ddmmyyyy or "",
            "MB_Brand": rig.get("MB_Brand", ""),
            "MB_Model_Raw": rig.get("MB_Model_Raw", ""),
            "CPU_Models": rig.get("CPU_Models", ""),
            "DDR": rig.get("DDR", ""),
            "RAM_Kits": rig.get("RAM_Kits", ""),
            "RAM_GB_Mentions": rig.get("RAM_GB_Mentions", ""),
            "Drive_Qty_Size_Mentions": rig.get("Drive_Qty_Size_Mentions", ""),
            "Platforms": rig.get("Platforms", ""),
            "Rig_Notes": rig_notes,
        })
    df_rigs = pd.DataFrame(rig_rows)

    # QA SAMPLE
    idle = df_features[df_features["System_Idle_Wall_Min_W"].notna()].copy()
    low_ids = idle.sort_values("System_Idle_Wall_Min_W").head(10)["Comment_ID"].tolist()
    high_ids = idle.sort_values("System_Idle_Wall_Min_W", ascending=False).head(10)["Comment_ID"].tolist()
    unclear_ids = df_power[df_power["Context_Class"] == "Unclear"].head(5)["Comment_ID"].tolist() if not df_power.empty else []

    sample_ids: List[int] = []
    for lst in (low_ids, high_ids, unclear_ids):
        for x in lst:
            if x not in sample_ids:
                sample_ids.append(x)

    df_qa = df_comments[df_comments["Comment_ID"].isin(sample_ids)].merge(df_features, on="Comment_ID", how="left")
    if not df_power.empty:
        df_qa = df_qa.merge(df_power.groupby("Comment_ID", as_index=False).agg(Watt_Mentions=("Watts", "count")), on="Comment_ID", how="left")
    else:
        df_qa["Watt_Mentions"] = 0

    df_qa["Reviewer_Override_Context_Class"] = ""
    df_qa["Reviewer_Override_Watts"] = ""
    df_qa["Reviewer_Notes"] = ""

    # INSIGHTS
    insights: Dict[str, pd.DataFrame] = {}

    if not df_power.empty:
        insights["Measurement_Context_Breakdown"] = (
            df_power.groupby(["Context_Class", "Confidence"], as_index=False)
            .agg(N=("Watts", "count"))
            .sort_values("N", ascending=False)
        )
    else:
        insights["Measurement_Context_Breakdown"] = pd.DataFrame(columns=["Context_Class", "Confidence", "N"])

    idle_wall_series = df_power[df_power["Context_Class"] == "System_Idle_Wall"]["Watts"] if not df_power.empty else pd.Series(dtype=float)
    if len(idle_wall_series):
        insights["Idle_Watts_Distribution_System_Idle_Wall"] = pd.DataFrame([{
            "N": int(len(idle_wall_series)),
            "Min": float(idle_wall_series.min()),
            "P10": float(idle_wall_series.quantile(0.10)),
            "Median": float(idle_wall_series.median()),
            "P90": float(idle_wall_series.quantile(0.90)),
            "Max": float(idle_wall_series.max()),
            "Mean": float(idle_wall_series.mean()),
        }])
    else:
        insights["Idle_Watts_Distribution_System_Idle_Wall"] = pd.DataFrame([{"N": 0, "Min": None, "P10": None, "Median": None, "P90": None, "Max": None, "Mean": None}])

    if not idle.empty:
        top_low = idle.sort_values("System_Idle_Wall_Min_W").head(25).merge(
            df_comments[["Comment_ID", "Commenter", "Date_DDMMYYYY"]], on="Comment_ID", how="left"
        )
        insights["Top_Low_Idle_Claims"] = top_low[["Comment_ID", "Commenter", "Date_DDMMYYYY", "System_Idle_Wall_Min_W", "Evidence_Score_1to5"]]
    else:
        insights["Top_Low_Idle_Claims"] = pd.DataFrame(columns=["Comment_ID", "Commenter", "Date_DDMMYYYY", "System_Idle_Wall_Min_W", "Evidence_Score_1to5"])

    if not df_themes.empty:
        insights["Theme_Frequency"] = (
            df_themes.groupby("Theme", as_index=False)
            .agg(Comments=("Comment_ID", "nunique"))
            .sort_values("Comments", ascending=False)
        )
    else:
        insights["Theme_Frequency"] = pd.DataFrame(columns=["Theme", "Comments"])

    return df_comments, df_rigs, df_power, df_themes, df_features, df_qa, insights


def build_readme(url: str) -> pd.DataFrame:
    return pd.DataFrame([
        {"Field": "Source URL", "Value": url},
        {"Field": "Comments", "Value": "One row per comment. Key fields: Commenter, Date, Threading, full comment text."},
        {"Field": "Rigs", "Value": "Best-effort rig/build extraction from comments (motherboard/CPU/RAM/storage/platform signals)."},
        {"Field": "Power_Observations", "Value": "One row per watt mention with sentence-bounded snippet + context classification (idle/load/psu/spec)."},
        {"Field": "Themes", "Value": "Theme tags per comment WITH the full comment text (pivot-friendly and reviewable)."},
        {"Field": "Insights", "Value": "Ready-to-read summary tables (use strict System_Idle_Wall for high-integrity stats)."},
        {"Field": "QA_Sample", "Value": "Shortlist to manually validate the most decision-sensitive claims."},
        {"Field": "Accuracy note", "Value": "Rig parsing and context classification are rule-based. Use QA_Sample to validate and refine patterns for your use-case."},
    ])


def write_workbook(out_path: str,
                   df_readme: pd.DataFrame,
                   df_comments: pd.DataFrame,
                   df_rigs: pd.DataFrame,
                   df_power: pd.DataFrame,
                   df_themes: pd.DataFrame,
                   df_features: pd.DataFrame,
                   df_qa: pd.DataFrame,
                   insights: Dict[str, pd.DataFrame]) -> None:
    for df in (df_readme, df_comments, df_rigs, df_power, df_themes, df_features, df_qa):
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].map(lambda x: excel_safe_text(x))

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df_readme.to_excel(w, sheet_name="README", index=False)
        df_comments.to_excel(w, sheet_name="Comments", index=False)
        df_rigs.to_excel(w, sheet_name="Rigs", index=False)
        df_power.to_excel(w, sheet_name="Power_Observations", index=False)
        df_themes.to_excel(w, sheet_name="Themes", index=False)
        df_features.to_excel(w, sheet_name="Comment_Features", index=False)
        df_qa.to_excel(w, sheet_name="QA_Sample", index=False)

        sheet = "Insights"
        row = 0
        for name, df in insights.items():
            pd.DataFrame({name: []}).to_excel(w, sheet_name=sheet, index=False, startrow=row)
            row += 1
            df.to_excel(w, sheet_name=sheet, index=False, startrow=row)
            row += len(df) + 3

    finalize_workbook(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", required=True, help="Article URL")
    ap.add_argument("--out", required=True, help="Output xlsx path")
    ap.add_argument("--timeout", type=int, default=30)
    args = ap.parse_args()

    html = fetch_html(args.url, timeout=args.timeout)
    soup = BeautifulSoup(html, "lxml")
    top_level = extract_top_level_comments(soup)
    records = walk_comments(top_level)

    if not records:
        raise SystemExit("No comments extracted. Markup may have changed.")

    df_readme = build_readme(args.url)
    df_comments, df_rigs, df_power, df_themes, df_features, df_qa, insights = build_tables(records)

    write_workbook(
        args.out,
        df_readme=df_readme,
        df_comments=df_comments,
        df_rigs=df_rigs,
        df_power=df_power,
        df_themes=df_themes,
        df_features=df_features,
        df_qa=df_qa,
        insights=insights,
    )

    n_comments = len(df_comments)
    n_power = len(df_power)
    n_idle_wall = int((df_power["Context_Class"] == "System_Idle_Wall").sum()) if n_power else 0
    n_rigs = len(df_rigs)
    print(f"OK: wrote {args.out} | comments={n_comments} | rigs={n_rigs} | watt_mentions={n_power} | system_idle_wall_mentions={n_idle_wall}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
